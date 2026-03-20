"""
Telegram-бот для учёта долгов в формате:
- /credit <число> <валюта> -> номер перевода (текст)
- /debit <номер карты/телефона> <банк> <сумма RUB> -> запрос в твой чат -> фото чека -> запись в базу
"""

import logging
import math
import os
import tempfile
from decimal import Decimal, ROUND_CEILING, ROUND_HALF_UP
from datetime import datetime

from aiogram import Bot, Dispatcher, F, Router
from aiogram.exceptions import TelegramBadRequest
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.fsm.state import State, StatesGroup
from aiogram.types import CallbackQuery, FSInputFile, Message
from aiogram.utils.keyboard import InlineKeyboardBuilder

from config import (
    BOT_TOKEN,
    ALLOWED_USER_IDS,
    PROXY,
    DEBT_BOT_MY_CHAT_ID,
    DEBT_BOT_PARTNER_CHAT_ID,
)
from database import (
    init_db,
    add_credit,
    add_debit_request,
    confirm_debit,
    reset_all,
    get_credits,
    get_debits_confirmed,
    get_debits_pending,
    get_total_credit_rub,
    get_total_debit_confirmed_rub,
    get_total_debt_rub,
)
from rate_parser import get_usdt_to_rub_rate

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)


def _create_bot() -> Bot:
    if PROXY:
        from aiogram.client.session.aiohttp import AiohttpSession

        session = AiohttpSession(proxy=PROXY)
        return Bot(token=BOT_TOKEN, session=session)
    return Bot(token=BOT_TOKEN)


bot = _create_bot()
router = Router()


def check_access(user_id: int) -> bool:
    if not ALLOWED_USER_IDS:
        return True
    return user_id in ALLOWED_USER_IDS


def fmt_sum(v: float) -> str:
    return f"{v:,.2f}".replace(",", " ").replace(".", ",")


def _balance_text(debt: float) -> str:
    if debt >= 0:
        return f"Долг: {fmt_sum(debt)} ₽"
    return f"Плюс: {fmt_sum(abs(debt))} ₽"


async def _safe_edit(message: Message, text: str, **kwargs):
    try:
        await message.edit_text(text, **kwargs)
    except TelegramBadRequest as e:
        # иногда Telegram кидает 400, если содержимое не изменилось
        if "message is not modified" not in str(e):
            raise


def build_menu() -> object:
    kb = InlineKeyboardBuilder()
    kb.button(text="📊 Excel", callback_data="export_excel")
    kb.button(text="🧹 Обнулить", callback_data="reset_start")
    kb.adjust(1)
    return kb.as_markup()


async def _show_main_screen(target_message: Message):
    """Обновить главное сообщение с кнопками."""
    debt = await get_total_debt_rub()
    await _safe_edit(
        target_message,
        f"💰 {_balance_text(debt)}\n\nКнопки ниже:",
        reply_markup=build_menu(),
    )


class CreditStates(StatesGroup):
    wait_transfer_number = State()


def _parse_credit_args(text: str) -> tuple[float, str] | None:
    parts = text.strip().split()
    # /credit <amount> <currency>
    if len(parts) < 3:
        return None
    amount_raw = parts[1].replace(",", ".")
    currency_raw = parts[2].strip().lower()

    try:
        amount = float(amount_raw)
        if amount <= 0:
            return None
    except ValueError:
        return None

    if currency_raw in {"usdt", "usd₮", "tether"}:
        currency = "USDT"
    elif currency_raw in {"rub", "rur", "₽", "ru"}:
        currency = "RUB"
    else:
        return None

    return amount, currency


def _parse_debit_args(text: str) -> tuple[str, str, float] | None:
    parts = text.strip().split()
    # /debit <card_or_phone> <bank...> <amount_rub>
    if len(parts) < 4:
        return None
    phone_or_card = parts[1]
    amount_raw = parts[-1].replace(",", ".")
    try:
        amount_rub = float(amount_raw)
    except ValueError:
        return None
    if amount_rub <= 0:
        return None
    bank = " ".join(parts[2:-1]).strip()
    if not bank:
        return None
    return phone_or_card, bank, amount_rub


@router.message(Command("start"))
async def cmd_start(message: Message, state: FSMContext):
    if message.chat.id != DEBT_BOT_MY_CHAT_ID:
        return
    if not check_access(message.from_user.id):
        await message.answer("⛔ Доступ запрещён.")
        return

    await state.clear()
    debt = await get_total_debt_rub()
    await message.answer(
        f"Привет.\n{_balance_text(debt)}.\n\nНажми кнопку для выгрузки Excel:",
        reply_markup=build_menu(),
    )


@router.message(Command("credit"))
async def cmd_credit(message: Message, state: FSMContext):
    if not check_access(message.from_user.id):
        await message.answer("⛔ Доступ запрещён.")
        return

    parsed = _parse_credit_args(message.text or "")
    if not parsed:
        await message.answer("Использование: /credit 276 USDT или /credit 1500 RUB")
        return

    amount_input, currency = parsed

    if currency == "USDT":
        await message.answer("⏳ Получаю курс USDT→RUB (округляю вверх до копеек)...")
        rate = await get_usdt_to_rub_rate()
        if rate is None:
            await message.answer("Не удалось получить курс. Попробуй позже или используй RUB.")
            return
        d_amount = Decimal(str(amount_input))
        d_rate = Decimal(str(rate))
        d_rub = (d_amount * d_rate).quantize(Decimal("0.01"), rounding=ROUND_CEILING)
        amount_rub = float(d_rub)
    else:
        d_amount = Decimal(str(amount_input))
        amount_rub = float(d_amount.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP))

    await state.clear()
    await state.update_data(
        credit_amount_input=amount_input,
        credit_currency=currency,
        credit_amount_rub=amount_rub,
        credit_chat_id=message.chat.id,
    )
    await state.set_state(CreditStates.wait_transfer_number)
    await message.answer("Дальше введи текстом номер перевода (любая строка).")


@router.message(CreditStates.wait_transfer_number, F.text)
async def credit_transfer_number(message: Message, state: FSMContext):
    data = await state.get_data()
    expected_chat_id = data.get("credit_chat_id")
    if expected_chat_id and message.chat.id != int(expected_chat_id):
        # игнорируем сообщения не из того чата, где была команда /credit
        return

    transfer_number = (message.text or "").strip()
    if not transfer_number:
        await message.answer("Номер перевода не может быть пустым. Введи текстом.")
        return

    credit_id = await add_credit(
        amount_input=float(data["credit_amount_input"]),
        currency=str(data["credit_currency"]),
        amount_rub=float(data["credit_amount_rub"]),
        transfer_number=transfer_number[:2000],
    )
    await state.clear()

    debt = await get_total_debt_rub()
    await message.answer(
        f"✅ Credit #{credit_id} сохранён.\n{_balance_text(debt)}.",
        reply_markup=build_menu(),
    )


@router.message(Command("debit"))
async def cmd_debit(message: Message):
    # debit вызывает товарищ в его чате
    if message.chat.id != DEBT_BOT_PARTNER_CHAT_ID:
        await message.answer("Команда /debit доступна только в чате товарища.")
        return

    parsed = _parse_debit_args(message.text or "")
    if not parsed:
        await message.answer("Пример: /debit +7999 Сбербанк 1500")
        return

    phone_or_card, bank, amount_rub = parsed
    debit_id = await add_debit_request(phone_or_card=phone_or_card, bank=bank, amount_rub=amount_rub)

    # идём в твой чат и просим фото
    await bot.send_message(
        DEBT_BOT_MY_CHAT_ID,
        "Товарищ сделал запрос оплаты.\n\n"
        f"Запрос #{debit_id}:\n"
        f"• Реквизит: {phone_or_card}\n"
        f"• Банк: {bank}\n"
        f"• Сумма: {fmt_sum(amount_rub)} ₽\n\n"
        "Пришли фото чека оплаты в этот чат.",
        reply_markup=build_menu(),
    )
    await message.answer(
        "✅ Запрос принят."
    )


@router.message(F.photo)
async def on_photo_from_my_chat(message: Message):
    if message.chat.id != DEBT_BOT_MY_CHAT_ID:
        return

    pending = await get_debits_pending(limit=1)
    if not pending:
        await message.answer("Фото принято, но сейчас нет ожидающих запросов /debit.")
        return

    debit = pending[0]
    photo_file_id = message.photo[-1].file_id
    ok = await confirm_debit(debit_id=int(debit["id"]), photo_file_id=photo_file_id)
    if not ok:
        await message.answer("Не удалось подтвердить оплату (запрос уже обработан).")
        return

    debt = await get_total_debt_rub()

    await message.answer(
        f"✅ Оплата подтверждена (debit #{debit['id']}).\n{_balance_text(debt)}."
    )
    await bot.send_message(
        DEBT_BOT_PARTNER_CHAT_ID,
        f"✅ Фото чека получено. Debit #{debit['id']} подтверждён.\n{_balance_text(debt)}.",
    )


@router.callback_query(F.data == "export_excel")
async def cb_export_excel(cb: CallbackQuery):
    if cb.message.chat.id != DEBT_BOT_MY_CHAT_ID:
        await cb.answer("Экспорт доступен только в твоём чате.", show_alert=True)
        return

    credits = await get_credits()
    debits_pending = await get_debits_pending(limit=1000)
    debits_confirmed = await get_debits_confirmed()
    total_credit = await get_total_credit_rub()
    total_debit_confirmed = await get_total_debit_confirmed_rub()
    debt = total_credit - total_debit_confirmed

    wb = Workbook()
    # Используем один лист
    ws_current = wb.active
    ws_current.title = "Текущие"

    header_font = Font(bold=True)
    align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)

    # Summary current
    ws_current.append(["Сальдо", _balance_text(debt)])
    ws_current.append([])
    ws_current.append(["Приход (credit) всего, ₽", float(total_credit)])
    ws_current.append(["Погашено (debit) подтверждено, ₽", float(total_debit_confirmed)])
    ws_current.append(["Активные debits (ждём фото), шт", int(len(debits_pending))])
    ws_current.append([])

    ws_current.append(["Кредиты"])
    ws_current.append(["id", "created_at", "amount_input", "currency", "amount_rub", "transfer_number"])
    for cell in ws_current[ws_current.max_row]:
        cell.font = header_font

    for c in credits:
        ws_current.append(
            [c["id"], c["created_at"], c["amount_input"], c["currency"], c["amount_rub"], c["transfer_number"]]
        )

    ws_current.append([])
    ws_current.append(["Debits pending (ждём фото)"])
    ws_current.append(["id", "created_at", "phone_or_card", "bank", "amount_rub"])
    for cell in ws_current[ws_current.max_row]:
        cell.font = header_font

    for d in debits_pending:
        ws_current.append([d["id"], d["created_at"], d["phone_or_card"], d["bank"], d["amount_rub"]])

    # Debits confirmed (погашено) — тоже на первой странице
    ws_current.append([])
    ws_current.append(["Debits confirmed (погашено)"])
    ws_current.append(["id", "created_at", "phone_or_card", "bank", "amount_rub", "photo_file_id"])
    for cell in ws_current[ws_current.max_row]:
        cell.font = header_font

    for d in debits_confirmed:
        ws_current.append(
            [d["id"], d["created_at"], d["phone_or_card"], d["bank"], d["amount_rub"], d["photo_file_id"]]
        )

    # autosize (примерно)
    for ws in (ws_current,):
        for col in range(1, ws.max_column + 1):
            width = 12
            for row in range(1, ws.max_row + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                width = max(width, min(60, len(str(v)) + 2))
            ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = align_left

    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp.close()
    # ВАЖНО: не удаляем файл сразу после отправки — иначе Telegram/aiogram
    # может не успеть дочитать содержимое, и Excel откроется битым.
    wb.save(tmp.name)
    try:
        size = os.path.getsize(tmp.name)
    except Exception:
        size = -1
    logger.info("Excel saved to %s (bytes=%s)", tmp.name, size)

    file = FSInputFile(tmp.name)
    await cb.message.answer_document(
        document=file,
        caption=f"Excel выгрузка.\n{_balance_text(debt)}",
    )

    await cb.answer()


@router.callback_query(F.data == "reset_start")
async def cb_reset_start(cb: CallbackQuery):
    if cb.message.chat.id != DEBT_BOT_MY_CHAT_ID:
        await cb.answer("Сброс доступен только в твоём чате.", show_alert=True)
        return
    if not check_access(cb.from_user.id):
        await cb.answer("⛔ Доступ запрещён", show_alert=True)
        return

    kb = InlineKeyboardBuilder()
    kb.button(text="✅ Да, обнулить", callback_data="reset_confirm_yes")
    kb.button(text="❌ Отмена", callback_data="reset_confirm_no")
    kb.adjust(1)

    await _safe_edit(
        cb.message,
        "Точно обнулить таблицу credit/debit?\n\nЭто удалит все записи.",
        reply_markup=kb.as_markup(),
    )
    await cb.answer()


@router.callback_query(F.data == "reset_confirm_no")
async def cb_reset_confirm_no(cb: CallbackQuery):
    if cb.message.chat.id != DEBT_BOT_MY_CHAT_ID:
        await cb.answer()
        return
    await _show_main_screen(cb.message)
    await cb.answer()


@router.callback_query(F.data == "reset_confirm_yes")
async def cb_reset_confirm_yes(cb: CallbackQuery):
    if cb.message.chat.id != DEBT_BOT_MY_CHAT_ID:
        await cb.answer()
        return
    if not check_access(cb.from_user.id):
        await cb.answer("⛔ Доступ запрещён", show_alert=True)
        return

    await reset_all()
    await _show_main_screen(cb.message)
    await cb.answer("✅ Обнулено")


async def main():
    await init_db()
    dp = Dispatcher()
    dp.include_router(router)
    await dp.start_polling(bot)


if __name__ == "__main__":
    import asyncio

    asyncio.run(main())
